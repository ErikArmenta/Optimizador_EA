# -*- coding: utf-8 -*-
"""
🏭 EA Simplex Production Optimizer v2.0
Motor de Optimización de Producción Industrial — LP + MILP + Restricciones Mixtas.
Desarrollado en Python por el Ingeniero Maestro Erik Armenta.
EA Innovation & Solutions — Ciudad Juárez, MX.

Mejoras v2.0:
  - MILP: variables enteras por producto
  - Restricciones mixtas: <=, >=, =
  - Importación desde Excel/CSV con plantilla descargable
  - Precios sombra (dual variables)
  - Comparador de escenarios
  - Análisis What-If interactivo
  - Persistencia de modelos en JSON
  - Validación de inputs y mensajes de error claros

"La exactitud es nuestra firma e innovar es nuestra naturaleza"
"""

# ============================================================
# IMPORTACIONES
# ============================================================
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import altair as alt
from scipy.optimize import linprog, milp, LinearConstraint, Bounds
from PIL import Image
from datetime import datetime
import io
import os
import json

try:
    from fpdf import FPDF
    PDF_ENABLED = True
except ImportError:
    PDF_ENABLED = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_ENABLED = True
except ImportError:
    EXCEL_ENABLED = False

# ============================================================
# CONFIGURACIÓN DE PÁGINA
# ============================================================
st.set_page_config(
    page_title="EA Simplex Optimizer v2",
    layout="wide",
    page_icon="🏭",
    initial_sidebar_state="expanded"
)

COLORS = {
    "bg_dark": "#0D0D0D",
    "bg_card": "#1A1A1A",
    "bg_card_hover": "#242424",
    "red_primary": "#C62828",
    "red_light": "#EF5350",
    "gold_primary": "#F9A825",
    "gold_light": "#FDD835",
    "white": "#FFFFFF",
    "text_primary": "#F5F5F5",
    "text_secondary": "#B0B0B0",
    "text_muted": "#707070",
    "border": "#333333",
    "success": "#00E676",
    "error": "#FF1744",
    "accent_blue": "#42A5F5",
}

# ============================================================
# CSS PREMIUM (mismo estilo + nuevas clases v2)
# ============================================================
st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500;700&display=swap');

    html {{ overflow-y: scroll; scroll-behavior: smooth; }}
    .main {{ background-color: {COLORS["bg_dark"]}; font-family: 'Inter', sans-serif; }}

    [data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #111111 0%, #1A1A1A 40%, #0D0D0D 100%);
        border-right: 1px solid {COLORS["border"]};
    }}
    [data-testid="stSidebar"] .stMarkdown h3 {{
        color: {COLORS["gold_primary"]} !important;
        font-weight: 700; letter-spacing: 1px; font-size: 14px; text-transform: uppercase;
    }}
    h1, h2, h3 {{ font-family: 'Inter', sans-serif !important; color: {COLORS["text_primary"]} !important; }}

    .glass-card {{
        background: rgba(26, 26, 26, 0.85);
        backdrop-filter: blur(16px);
        border: 1px solid rgba(255, 255, 255, 0.06);
        border-radius: 16px; padding: 28px; margin-bottom: 20px;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.3);
    }}
    .glass-card:hover {{
        border-color: rgba(249, 168, 37, 0.2);
        box-shadow: 0 12px 48px rgba(249, 168, 37, 0.08);
        transform: translateY(-2px);
    }}

    .hero-header {{
        background: linear-gradient(135deg, #0D0D0D 0%, #1A1A1A 50%, #0D0D0D 100%);
        border: 1px solid {COLORS["border"]}; border-radius: 20px;
        padding: 30px 40px; margin-bottom: 24px;
        position: relative; overflow: hidden;
    }}
    .hero-header::before {{
        content: ''; position: absolute; top: 0; left: 0; right: 0; height: 3px;
        background: linear-gradient(90deg, {COLORS["red_primary"]}, {COLORS["gold_primary"]}, {COLORS["red_primary"]});
        animation: shimmer 3s ease-in-out infinite;
    }}
    @keyframes shimmer {{ 0%, 100% {{ opacity: 0.6; }} 50% {{ opacity: 1; }} }}
    .hero-title {{
        font-size: 2.2rem; font-weight: 900;
        background: linear-gradient(135deg, {COLORS["white"]} 0%, {COLORS["gold_primary"]} 100%);
        -webkit-background-clip: text; -webkit-text-fill-color: transparent;
        background-clip: text; margin: 0; line-height: 1.2;
    }}
    .hero-subtitle {{
        color: {COLORS["text_secondary"]}; font-size: 1rem; font-weight: 400;
        letter-spacing: 3px; text-transform: uppercase; margin-top: 6px;
    }}
    .hero-slogan {{
        color: {COLORS["gold_primary"]}; font-size: 0.85rem;
        font-style: italic; font-weight: 500; margin-top: 10px; opacity: 0.85;
    }}

    .kpi-container {{
        display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr));
        gap: 16px; margin: 20px 0;
    }}
    .kpi-card {{
        background: linear-gradient(145deg, #1A1A1A, #222222);
        border: 1px solid {COLORS["border"]}; border-radius: 14px;
        padding: 22px; text-align: center; position: relative;
        overflow: hidden; transition: all 0.3s ease;
    }}
    .kpi-card:hover {{ transform: translateY(-4px); box-shadow: 0 12px 40px rgba(0,0,0,0.4); }}
    .kpi-card::after {{ content: ''; position: absolute; bottom: 0; left: 0; right: 0; height: 3px; }}
    .kpi-card.red::after {{ background: {COLORS["red_primary"]}; }}
    .kpi-card.gold::after {{ background: {COLORS["gold_primary"]}; }}
    .kpi-card.success::after {{ background: {COLORS["success"]}; }}
    .kpi-card.blue::after {{ background: {COLORS["accent_blue"]}; }}
    .kpi-label {{ font-size: 0.78rem; color: {COLORS["text_secondary"]}; text-transform: uppercase; letter-spacing: 1.5px; font-weight: 600; margin-bottom: 8px; }}
    .kpi-value {{ font-size: 1.8rem; font-weight: 800; color: {COLORS["white"]}; font-family: 'JetBrains Mono', monospace; }}
    .kpi-unit {{ font-size: 0.8rem; color: {COLORS["text_muted"]}; margin-top: 4px; }}

    .status-badge {{
        display: inline-block; padding: 4px 12px; border-radius: 20px;
        font-size: 0.75rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.5px;
    }}
    .badge-holgura {{ background: rgba(0,230,118,0.15); color: {COLORS["success"]}; }}
    .badge-activa {{ background: rgba(198,40,40,0.15); color: {COLORS["red_light"]}; }}
    .badge-max {{ background: rgba(249,168,37,0.15); color: {COLORS["gold_primary"]}; }}
    .badge-min {{ background: rgba(66,165,245,0.15); color: {COLORS["accent_blue"]}; }}
    .badge-int {{ background: rgba(171,71,188,0.15); color: #CE93D8; }}
    .badge-cont {{ background: rgba(66,165,245,0.1); color: {COLORS["accent_blue"]}; }}

    .stButton > button {{
        background: linear-gradient(135deg, {COLORS["red_primary"]}, #D32F2F) !important;
        color: white !important; font-weight: 700 !important; border: none !important;
        border-radius: 10px !important; padding: 12px 28px !important;
        font-family: 'Inter', sans-serif !important;
        text-transform: uppercase !important; letter-spacing: 1px !important;
        transition: all 0.3s ease !important;
        box-shadow: 0 4px 15px rgba(198,40,40,0.3) !important;
    }}
    .stButton > button:hover {{ transform: translateY(-2px) !important; box-shadow: 0 8px 25px rgba(198,40,40,0.45) !important; }}

    .stDownloadButton > button {{
        background: linear-gradient(135deg, {COLORS["gold_primary"]}, #F9A825) !important;
        color: #0D0D0D !important; font-weight: 700 !important; border: none !important;
        border-radius: 10px !important; letter-spacing: 1px !important;
        text-transform: uppercase !important;
        box-shadow: 0 4px 15px rgba(249,168,37,0.3) !important;
    }}
    .stDownloadButton > button:hover {{ transform: translateY(-2px) !important; box-shadow: 0 8px 25px rgba(249,168,37,0.45) !important; }}

    .stNumberInput input, .stTextInput input {{
        background-color: #1A1A1A !important; border: 1px solid {COLORS["border"]} !important;
        border-radius: 8px !important; color: {COLORS["text_primary"]} !important;
        font-family: 'JetBrains Mono', monospace !important;
    }}
    .stNumberInput input:focus, .stTextInput input:focus {{
        border-color: {COLORS["gold_primary"]} !important;
        box-shadow: 0 0 0 2px rgba(249,168,37,0.2) !important;
    }}
    .stSelectbox > div > div {{
        background-color: #1A1A1A !important;
        border: 1px solid {COLORS["border"]} !important; border-radius: 8px !important;
    }}
    .streamlit-expanderHeader {{
        background-color: {COLORS["bg_card"]} !important; border-radius: 10px !important;
        border: 1px solid {COLORS["border"]} !important;
        color: {COLORS["text_primary"]} !important; font-weight: 600 !important;
    }}
    .stTabs [data-baseweb="tab-list"] {{ gap: 8px; }}
    .stTabs [data-baseweb="tab"] {{
        background: {COLORS["bg_card"]} !important;
        border-radius: 10px 10px 0 0 !important;
        border: 1px solid {COLORS["border"]} !important;
        color: {COLORS["text_secondary"]} !important; font-weight: 600 !important;
        padding: 10px 20px !important;
    }}
    .stTabs [aria-selected="true"] {{
        background: linear-gradient(135deg, {COLORS["red_primary"]}, #D32F2F) !important;
        color: white !important; border-color: {COLORS["red_primary"]} !important;
    }}
    .footer-premium {{
        background: linear-gradient(135deg, #111111, #1A1A1A);
        border-top: 1px solid {COLORS["border"]}; border-radius: 16px;
        padding: 24px 32px; margin-top: 40px; text-align: center;
    }}
    .footer-slogan {{ color: {COLORS["gold_primary"]}; font-size: 0.95rem; font-weight: 600; font-style: italic; letter-spacing: 1.5px; }}
    .footer-credits {{ color: {COLORS["text_muted"]}; font-size: 0.78rem; margin-top: 8px; }}
    ::-webkit-scrollbar {{ width: 8px; }}
    ::-webkit-scrollbar-track {{ background: {COLORS["bg_dark"]}; }}
    ::-webkit-scrollbar-thumb {{ background: {COLORS["border"]}; border-radius: 4px; }}
    ::-webkit-scrollbar-thumb:hover {{ background: {COLORS["text_muted"]}; }}
    @keyframes fadeInUp {{ from {{ opacity: 0; transform: translateY(20px); }} to {{ opacity: 1; transform: translateY(0); }} }}
    .animate-in {{ animation: fadeInUp 0.6s ease-out forwards; }}
    .section-divider {{ height: 1px; background: linear-gradient(90deg, transparent, {COLORS["border"]}, transparent); margin: 30px 0; }}
    .pulse-icon {{ display: inline-block; animation: pulse 2s ease-in-out infinite; }}
    @keyframes pulse {{ 0%, 100% {{ transform: scale(1); }} 50% {{ transform: scale(1.15); }} }}
    /* v2: badge verde para MILP */
    .v2-badge {{
        background: linear-gradient(135deg, rgba(198,40,40,0.2), rgba(249,168,37,0.2));
        border: 1px solid rgba(249,168,37,0.4); border-radius: 20px;
        padding: 3px 10px; font-size: 0.7rem; font-weight: 700;
        color: {COLORS["gold_primary"]}; text-transform: uppercase; letter-spacing: 1px;
    }}
    #MainMenu {{ visibility: hidden; }} footer {{ visibility: hidden; }} header {{ visibility: hidden; }}
</style>
""", unsafe_allow_html=True)


# ============================================================
# UTILIDADES
# ============================================================

def sanitize_text(text: str) -> str:
    replacements = {
        '\u2014': ' - ', '\u2013': ' - ', '\u2022': '*', '\u2019': "'",
        '\u201c': '"', '\u201d': '"', '\u2018': "'", '\u2026': '...',
        'á': 'a', 'é': 'e', 'í': 'i', 'ó': 'o', 'ú': 'u',
        'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U',
        'ñ': 'n', 'Ñ': 'N', '¿': '?', '¡': '!', '≤': '<=', '≥': '>=',
    }
    for uni, asc in replacements.items():
        text = text.replace(uni, asc)
    return text.encode('latin-1', errors='replace').decode('latin-1')


STATUS_MESSAGES = {
    0: "Solución óptima encontrada.",
    1: "Límite de iteraciones alcanzado. El problema puede ser no acotado.",
    2: "El problema es INFACTIBLE. Las restricciones se contradicen entre sí. "
       "Verifica que no haya restricciones ≥ y ≤ que se excluyan mutuamente.",
    3: "El problema es NO ACOTADO. La función objetivo puede crecer sin límite. "
       "Agrega restricciones de demanda máxima o límites de producción.",
    4: "Problemas numéricos. Revisa coeficientes muy grandes o muy pequeños.",
}


@st.cache_data(show_spinner=False)
def resolver_problema(c_tuple, A_tuple, b_tuple, tipos_tuple, enteros_tuple, objetivo):
    """
    Resuelve LP/MILP con restricciones mixtas usando scipy.optimize.milp.
    Devuelve precios sombra via relajación LP con linprog (solo LP puro).
    """
    c = list(c_tuple)
    A = [list(row) for row in A_tuple]
    b = list(b_tuple)
    tipos = list(tipos_tuple)
    es_entero = list(enteros_tuple)
    n = len(c)

    c_opt = np.array([-ci for ci in c] if objetivo == "Maximizar" else c, dtype=float)
    A_np = np.array(A, dtype=float)
    b_np = np.array(b, dtype=float)

    # Construir bounds de restricciones para milp
    lb_con, ub_con = [], []
    for tipo, bi in zip(tipos, b):
        if tipo == "≤":
            lb_con.append(-np.inf); ub_con.append(bi)
        elif tipo == "≥":
            lb_con.append(bi); ub_con.append(np.inf)
        else:  # "="
            lb_con.append(bi); ub_con.append(bi)

    constraints = LinearConstraint(A_np, lb_con, ub_con)
    integrality = np.array([1.0 if e else 0.0 for e in es_entero])
    var_bounds = Bounds(lb=0.0, ub=np.inf)

    res_milp = milp(c_opt, constraints=constraints, integrality=integrality, bounds=var_bounds)

    if res_milp.status != 0:
        msg = STATUS_MESSAGES.get(res_milp.status, res_milp.message)
        return {"success": False, "status": res_milp.status, "mensaje": msg}

    variables = res_milp.x
    valor_obj = -res_milp.fun if objetivo == "Maximizar" else res_milp.fun
    consumo = A_np @ variables

    # Holguras (positivo = sobra; para >= es cuánto supera el mínimo)
    holguras = []
    for j, (tipo, bi) in enumerate(zip(tipos, b)):
        if tipo == "≤":
            holguras.append(bi - consumo[j])
        elif tipo == "≥":
            holguras.append(consumo[j] - bi)
        else:
            holguras.append(0.0)
    holguras = np.array(holguras)

    # ---- Precios sombra via LP relajación (solo si no hay enteros) ----
    duals = np.zeros(len(b))
    tiene_enteros = any(es_entero)
    if not tiene_enteros:
        A_ub_lp, b_ub_lp, A_eq_lp, b_eq_lp = [], [], [], []
        map_ub, map_eq = [], []  # índice original de cada restricción

        for j, (tipo, bi, row) in enumerate(zip(tipos, b, A)):
            if tipo == "≤":
                A_ub_lp.append(row); b_ub_lp.append(bi); map_ub.append(j)
            elif tipo == "≥":
                A_ub_lp.append([-x for x in row]); b_ub_lp.append(-bi); map_ub.append(j)
            else:
                A_eq_lp.append(row); b_eq_lp.append(bi); map_eq.append(j)

        lp_kwargs = {"c": c_opt, "bounds": [(0, None)] * n, "method": "highs"}
        if A_ub_lp:
            lp_kwargs["A_ub"] = np.array(A_ub_lp)
            lp_kwargs["b_ub"] = np.array(b_ub_lp)
        if A_eq_lp:
            lp_kwargs["A_eq"] = np.array(A_eq_lp)
            lp_kwargs["b_eq"] = np.array(b_eq_lp)

        lp_res = linprog(**lp_kwargs)
        if lp_res.success:
            ub_duals = (lp_res.ineqlin.marginals
                        if hasattr(lp_res, "ineqlin") and lp_res.ineqlin is not None
                        else np.zeros(len(A_ub_lp)))
            eq_duals = (lp_res.eqlin.marginals
                        if hasattr(lp_res, "eqlin") and lp_res.eqlin is not None
                        else np.zeros(len(A_eq_lp)))

            for k, orig_j in enumerate(map_ub):
                raw = ub_duals[k] if k < len(ub_duals) else 0.0
                if tipos[orig_j] == "≥":
                    raw = -raw
                duals[orig_j] = -raw if objetivo == "Maximizar" else raw

            for k, orig_j in enumerate(map_eq):
                raw = eq_duals[k] if k < len(eq_duals) else 0.0
                duals[orig_j] = -raw if objetivo == "Maximizar" else raw

    return {
        "success": True,
        "variables": variables.tolist(),
        "valor_objetivo": float(valor_obj),
        "holguras": holguras.tolist(),
        "consumo": consumo.tolist(),
        "duals": duals.tolist(),
        "tiene_enteros": tiene_enteros,
    }


def validar_inputs(margenes, coefs, nombres_prod, nombres_rest):
    """Retorna lista de advertencias de validación."""
    warnings = []
    for i, (m, n) in enumerate(zip(margenes, nombres_prod)):
        if m == 0:
            warnings.append(f"⚠️ El producto **{n}** tiene utilidad/costo = 0. No impacta el objetivo.")
    for j, (fila, nombre) in enumerate(zip(coefs, nombres_rest)):
        if all(c == 0 for c in fila):
            warnings.append(f"⚠️ La restricción **{nombre}** tiene todos los coeficientes en 0 — nunca restringe nada.")
    for i, n in enumerate(zip(nombres_prod, range(len(margenes)))):
        col_i = [coefs[j][i] for j in range(len(coefs))]
        if all(c == 0 for c in col_i):
            warnings.append(f"⚠️ El producto **{nombres_prod[i]}** tiene coeficiente 0 en todas las restricciones — es ilimitado.")
    return warnings


def generar_plantilla_excel() -> bytes:
    """Genera una plantilla Excel descargable para importar datos."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Modelo"

    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")

    # Instrucciones
    ws["A1"] = "EA Simplex Optimizer v2 — Plantilla de Importación"
    ws["A1"].font = Font(name="Arial", bold=True, size=13, color="C62828")
    ws["A2"] = "Completa la tabla desde la fila 5. No modifiques los encabezados."
    ws["A2"].font = Font(name="Arial", size=10, color="888888", italic=True)

    # Encabezados desde fila 4
    headers = ["Restriccion", "Tipo (<=, >=, =)", "Limite", "Producto1", "Producto2", "Producto3"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # Datos de ejemplo
    ejemplo = [
        ["Horas Maquina", "<=", 100, 2, 3, 1],
        ["Materia Prima", "<=", 150, 4, 1, 2],
        ["Demanda Min Prod1", ">=", 10, 1, 0, 0],
    ]
    for i, row in enumerate(ejemplo, start=5):
        for j, val in enumerate(row, start=1):
            ws.cell(row=i, column=j, value=val)

    # Hoja de productos
    ws2 = wb.create_sheet("Productos")
    ws2["A1"] = "Nombre"
    ws2["B1"] = "Utilidad_Costo"
    ws2["C1"] = "Es_Entero (0 o 1)"
    ws2["A1"].font = header_font; ws2["A1"].fill = header_fill
    ws2["B1"].font = header_font; ws2["B1"].fill = header_fill
    ws2["C1"].font = header_font; ws2["C1"].fill = header_fill
    for i, (n, u, e) in enumerate([("Producto1", 10, 0), ("Producto2", 15, 0), ("Producto3", 8, 1)], start=2):
        ws2.cell(row=i, column=1, value=n)
        ws2.cell(row=i, column=2, value=u)
        ws2.cell(row=i, column=3, value=e)

    for ws_tmp in [ws, ws2]:
        for col in ws_tmp.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws_tmp.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def leer_excel_importado(file) -> dict | None:
    """Lee el Excel de importación y retorna dict con datos del modelo."""
    try:
        df_prod = pd.read_excel(file, sheet_name="Productos", header=0)
        df_rest = pd.read_excel(file, sheet_name="Modelo", header=3)  # fila 4 = índice 3

        nombres_prod = df_prod.iloc[:, 0].dropna().astype(str).tolist()
        margenes = df_prod.iloc[:, 1].fillna(0).astype(float).tolist()
        enteros = df_prod.iloc[:, 2].fillna(0).astype(int).tolist() if df_prod.shape[1] > 2 else [0] * len(nombres_prod)

        nombres_rest = df_rest.iloc[:, 0].dropna().astype(str).tolist()
        tipos = df_rest.iloc[:, 1].fillna("<=").astype(str).tolist()
        # Normalizar tipos
        tipo_map = {"<=": "≤", ">=": "≥", "=": "=", "≤": "≤", "≥": "≥"}
        tipos = [tipo_map.get(t.strip(), "≤") for t in tipos]
        limites = df_rest.iloc[:, 2].fillna(0).astype(float).tolist()

        n_prod = len(nombres_prod)
        coefs = []
        for _, row_data in df_rest.iterrows():
            fila = []
            for k in range(n_prod):
                val = row_data.iloc[3 + k] if 3 + k < len(row_data) else 0
                fila.append(float(val) if pd.notna(val) else 0.0)
            coefs.append(fila)
        coefs = coefs[:len(nombres_rest)]

        return {
            "nombres_productos": nombres_prod,
            "margenes": margenes,
            "es_entero": [bool(e) for e in enteros],
            "nombres_restricciones": nombres_rest,
            "tipos": tipos,
            "limites": limites,
            "coeficientes": coefs,
        }
    except Exception as e:
        st.error(f"Error al leer el archivo: {e}")
        return None


def generar_pdf_reporte(datos_reporte: dict) -> bytes:
    if not PDF_ENABLED:
        return None

    class SimplexPDF(FPDF):
        def header(self):
            self.set_fill_color(198, 40, 40)
            self.rect(0, 0, 210, 8, 'F')
            self.set_fill_color(249, 168, 37)
            self.rect(0, 8, 210, 2, 'F')
            self.set_y(15)
            self.set_font('Helvetica', 'B', 20)
            self.set_text_color(198, 40, 40)
            self.cell(0, 10, 'EA SIMPLEX OPTIMIZER v2', align='C', new_x="LMARGIN", new_y="NEXT")
            self.set_font('Helvetica', '', 10)
            self.set_text_color(120, 120, 120)
            self.cell(0, 6, 'Reporte Ejecutivo de Optimizacion de Produccion', align='C', new_x="LMARGIN", new_y="NEXT")
            self.ln(4)
            self.set_draw_color(198, 40, 40); self.set_line_width(0.5)
            self.line(15, self.get_y(), 195, self.get_y()); self.ln(6)

        def footer(self):
            self.set_y(-20)
            self.set_draw_color(249, 168, 37); self.set_line_width(0.3)
            self.line(15, self.get_y(), 195, self.get_y()); self.ln(3)
            self.set_font('Helvetica', 'I', 7); self.set_text_color(150, 150, 150)
            self.cell(0, 5, sanitize_text('"La exactitud es nuestra firma e innovar es nuestra naturaleza"'), align='C', new_x="LMARGIN", new_y="NEXT")
            self.cell(0, 5, f'Pagina {self.page_no()} | EA Innovation & Solutions | {datetime.now().strftime("%d/%m/%Y %H:%M")}', align='C')

    pdf = SimplexPDF()
    pdf.add_page()

    # === INFO GENERAL ===
    pdf.set_font('Helvetica', 'B', 13); pdf.set_text_color(198, 40, 40)
    pdf.cell(0, 8, '1. INFORMACION GENERAL', new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    pdf.set_font('Helvetica', '', 10); pdf.set_text_color(50, 50, 50)
    milp_flag = " (MILP - Variables Enteras)" if datos_reporte.get("tiene_enteros") else " (LP Continuo)"
    for line in [
        f"Objetivo: {datos_reporte['objetivo']}",
        f"Modo de solucion: {milp_flag}",
        f"Productos: {datos_reporte['num_productos']} | Restricciones: {datos_reporte['num_restricciones']}",
        f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}",
        f"Motor: SciPy MILP / HiGHS",
    ]:
        pdf.cell(0, 6, sanitize_text(line), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # === RESULTADO ÓPTIMO ===
    pdf.set_font('Helvetica', 'B', 13); pdf.set_text_color(198, 40, 40)
    pdf.cell(0, 8, '2. RESULTADO OPTIMO', new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    pdf.set_font('Helvetica', 'B', 24); pdf.set_text_color(249, 168, 37)
    pdf.cell(0, 14, sanitize_text(f"${datos_reporte['valor_objetivo']:,.2f}"), align='C', new_x="LMARGIN", new_y="NEXT")
    pdf.set_font('Helvetica', '', 9); pdf.set_text_color(120, 120, 120)
    label_obj = "Utilidad Maxima" if "Maximizar" in datos_reporte['objetivo'] else "Costo Minimo"
    pdf.cell(0, 6, sanitize_text(label_obj), align='C', new_x="LMARGIN", new_y="NEXT"); pdf.ln(6)

    # Tabla productos
    pdf.set_font('Helvetica', 'B', 11); pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 7, 'Plan de Produccion Optimo:', new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    pdf.set_fill_color(198, 40, 40); pdf.set_text_color(255, 255, 255); pdf.set_font('Helvetica', 'B', 9)
    col_w = [55, 40, 40, 30]
    for h, w in zip(['Producto', 'Cantidad Optima', 'Contribucion ($)', 'Tipo'], col_w):
        pdf.cell(w, 8, sanitize_text(h), border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(50, 50, 50); pdf.set_font('Helvetica', '', 9)
    for i, prod in enumerate(datos_reporte['productos']):
        fill = i % 2 == 0
        pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
        pdf.cell(col_w[0], 7, sanitize_text(prod['nombre']), border=1, fill=fill)
        pdf.cell(col_w[1], 7, f"{prod['cantidad']:.2f}", border=1, fill=fill, align='C')
        pdf.cell(col_w[2], 7, f"${prod['contribucion']:,.2f}", border=1, fill=fill, align='C')
        pdf.cell(col_w[3], 7, "Entero" if prod.get('es_entero') else "Continuo", border=1, fill=fill, align='C')
        pdf.ln()
    pdf.ln(6)

    # === ANÁLISIS DE SENSIBILIDAD ===
    pdf.set_font('Helvetica', 'B', 13); pdf.set_text_color(198, 40, 40)
    pdf.cell(0, 8, '3. ANALISIS DE SENSIBILIDAD', new_x="LMARGIN", new_y="NEXT"); pdf.ln(2)
    pdf.set_fill_color(198, 40, 40); pdf.set_text_color(255, 255, 255); pdf.set_font('Helvetica', 'B', 9)
    r_col_w = [45, 25, 25, 25, 25, 20]
    r_heads = ['Recurso', 'Disp.', 'Consumido', 'Holgura', 'P. Sombra', 'Estado']
    for h, w in zip(r_heads, r_col_w):
        pdf.cell(w, 8, sanitize_text(h), border=1, fill=True, align='C')
    pdf.ln()
    pdf.set_text_color(50, 50, 50); pdf.set_font('Helvetica', '', 9)
    for i, rec in enumerate(datos_reporte['recursos']):
        fill = i % 2 == 0
        pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
        estado = "ACTIVA" if rec['holgura'] < 0.01 else "HOLGURA"
        pdf.cell(r_col_w[0], 7, sanitize_text(rec['nombre']), border=1, fill=fill)
        pdf.cell(r_col_w[1], 7, f"{rec['disponible']:,.1f}", border=1, fill=fill, align='C')
        pdf.cell(r_col_w[2], 7, f"{rec['consumido']:,.1f}", border=1, fill=fill, align='C')
        pdf.cell(r_col_w[3], 7, f"{rec['holgura']:,.1f}", border=1, fill=fill, align='C')
        dual_str = f"{rec.get('dual', 0):,.2f}" if not datos_reporte.get('tiene_enteros') else "N/A"
        pdf.cell(r_col_w[4], 7, dual_str, border=1, fill=fill, align='C')
        pdf.cell(r_col_w[5], 7, estado, border=1, fill=fill, align='C')
        pdf.ln()

    if datos_reporte.get('tiene_enteros'):
        pdf.ln(4); pdf.set_font('Helvetica', 'I', 8); pdf.set_text_color(150, 150, 150)
        pdf.multi_cell(0, 5, sanitize_text(
            "Nota: Los precios sombra no aplican a MILP. "
            "Para obtenerlos, desactiva las variables enteras y resuelve la relajacion LP."
        ))
    else:
        pdf.ln(4); pdf.set_font('Helvetica', 'I', 8); pdf.set_text_color(150, 150, 150)
        pdf.multi_cell(0, 5, sanitize_text(
            "Precio Sombra: incremento en el valor objetivo por cada unidad adicional del recurso. "
            "Valor positivo en maximizacion = el recurso vale esa cantidad adicional si se amplia."
        ))

    return bytes(pdf.output())


def generar_excel_reporte(datos_reporte: dict) -> bytes:
    if not EXCEL_ENABLED:
        return None
    wb = openpyxl.Workbook()
    rojo_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    gris_claro = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
    blanco_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    fuente_titulo = Font(name='Arial', bold=True, size=14, color="C62828")
    fuente_header = Font(name='Arial', bold=True, size=10, color="FFFFFF")
    fuente_datos = Font(name='Arial', size=10, color="333333")
    fuente_kpi = Font(name='Arial', bold=True, size=18, color="F9A825")
    center = Alignment(horizontal='center', vertical='center')
    borde = Border(
        left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'),
        top=Side(style='thin', color='DDDDDD'), bottom=Side(style='thin', color='DDDDDD')
    )

    ws1 = wb.active; ws1.title = "Resumen Ejecutivo"; ws1.sheet_properties.tabColor = "C62828"
    for col, w in zip(['A','B','C','D','E'], [5, 30, 25, 25, 20]):
        ws1.column_dimensions[col].width = w

    ws1.merge_cells('B2:E2'); cell = ws1['B2']
    cell.value = "EA SIMPLEX OPTIMIZER v2 — Reporte Ejecutivo"
    cell.font = fuente_titulo; cell.alignment = center

    ws1.merge_cells('B3:E3'); cell = ws1['B3']
    cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Motor: SciPy MILP/HiGHS"
    cell.font = Font(name='Arial', size=9, color="999999", italic=True); cell.alignment = center

    ws1.merge_cells('B5:E5'); cell = ws1['B5']
    cell.value = "UTILIDAD MÁXIMA" if "Maximizar" in datos_reporte['objetivo'] else "COSTO MÍNIMO"
    cell.font = Font(name='Arial', bold=True, size=11, color="C62828"); cell.alignment = center

    ws1.merge_cells('B6:E6'); cell = ws1['B6']
    cell.value = f"${datos_reporte['valor_objetivo']:,.2f}"
    cell.font = fuente_kpi; cell.alignment = center
    cell.fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")

    row = 9
    ws1.cell(row=row, column=2, value="PLAN DE PRODUCCIÓN ÓPTIMO").font = Font(name='Arial', bold=True, size=11, color="C62828")
    row += 1
    for j, h in enumerate(["Producto", "Cantidad Óptima", "Contribución ($)", "Tipo Variable"], start=2):
        c = ws1.cell(row=row, column=j, value=h); c.font = fuente_header; c.fill = rojo_fill; c.alignment = center; c.border = borde
    row += 1
    for i, prod in enumerate(datos_reporte['productos']):
        fill = gris_claro if i % 2 == 0 else blanco_fill
        ws1.cell(row=row, column=2, value=prod['nombre']).font = fuente_datos; ws1.cell(row=row, column=2).fill = fill; ws1.cell(row=row, column=2).border = borde
        ws1.cell(row=row, column=3, value=round(prod['cantidad'], 2)).font = fuente_datos; ws1.cell(row=row, column=3).fill = fill; ws1.cell(row=row, column=3).alignment = center; ws1.cell(row=row, column=3).border = borde
        ws1.cell(row=row, column=4, value=round(prod['contribucion'], 2)).font = fuente_datos; ws1.cell(row=row, column=4).fill = fill; ws1.cell(row=row, column=4).alignment = center; ws1.cell(row=row, column=4).border = borde; ws1.cell(row=row, column=4).number_format = '$#,##0.00'
        ws1.cell(row=row, column=5, value="Entero" if prod.get('es_entero') else "Continuo").font = fuente_datos; ws1.cell(row=row, column=5).fill = fill; ws1.cell(row=row, column=5).alignment = center; ws1.cell(row=row, column=5).border = borde
        row += 1

    ws2 = wb.create_sheet("Analisis Sensibilidad"); ws2.sheet_properties.tabColor = "F9A825"
    for col, w in zip(['A','B','C','D','E','F','G','H'], [5, 28, 15, 15, 15, 18, 18, 18]):
        ws2.column_dimensions[col].width = w

    ws2.merge_cells('B2:H2'); cell = ws2['B2']
    cell.value = "ANÁLISIS DE SENSIBILIDAD DE RECURSOS v2"
    cell.font = fuente_titulo; cell.alignment = center

    row = 4
    for j, h in enumerate(["Recurso", "Tipo", "Disponible", "Consumido", "Holgura", "% Util.", "Precio Sombra", "Estado"], start=2):
        c = ws2.cell(row=row, column=j, value=h); c.font = fuente_header; c.fill = rojo_fill; c.alignment = center; c.border = borde
    row += 1
    for i, rec in enumerate(datos_reporte['recursos']):
        fill = gris_claro if i % 2 == 0 else blanco_fill
        pct = (rec['consumido'] / rec['disponible'] * 100) if rec['disponible'] > 0 else 0
        estado = "ACTIVA" if rec['holgura'] < 0.01 else "HOLGURA"
        dual_val = rec.get('dual', 0) if not datos_reporte.get('tiene_enteros') else None
        dual_str = f"{dual_val:.4f}" if dual_val is not None else "N/A (MILP)"
        vals = [rec['nombre'], rec.get('tipo', '≤'), round(rec['disponible'], 2), round(rec['consumido'], 2), round(rec['holgura'], 2), round(pct, 1), dual_str, estado]
        for j, v in enumerate(vals, start=2):
            c = ws2.cell(row=row, column=j, value=v); c.font = fuente_datos; c.fill = fill; c.alignment = center; c.border = borde
        ws2.cell(row=row, column=9).font = Font(name='Arial', bold=True, size=10, color="C62828" if estado == "ACTIVA" else "388E3C")
        row += 1

    if datos_reporte.get('escenarios'):
        ws3 = wb.create_sheet("Comparador Escenarios"); ws3.sheet_properties.tabColor = "42A5F5"
        ws3["B2"] = "COMPARADOR DE ESCENARIOS"; ws3["B2"].font = fuente_titulo
        row = 4
        for j, h in enumerate(["Escenario", "Objetivo", "Valor Obj.", "Modo"], start=2):
            c = ws3.cell(row=row, column=j, value=h); c.font = fuente_header; c.fill = rojo_fill; c.alignment = center; c.border = borde
        row += 1
        for i, esc in enumerate(datos_reporte['escenarios']):
            fill = gris_claro if i % 2 == 0 else blanco_fill
            for j, v in enumerate([esc['nombre'], esc['tipo_objetivo'], f"${esc['valor_objetivo']:,.2f}", "MILP" if esc.get('tiene_enteros') else "LP"], start=2):
                c = ws3.cell(row=row, column=j, value=v); c.font = fuente_datos; c.fill = fill; c.alignment = center; c.border = borde
            row += 1

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()


# ============================================================
# HEADER
# ============================================================
with st.container():
    col_logo, col_titulo = st.columns([1, 7])
    with col_logo:
        try:
            logo = Image.open("EA_2.png")
            st.image(logo, use_container_width=True)
        except Exception:
            st.markdown("<div style='font-size:50px;text-align:center;'>🏭</div>", unsafe_allow_html=True)
    with col_titulo:
        st.markdown("""
        <div class="hero-header">
            <p class="hero-title">🏭 Simplex Production Optimizer</p>
            <p class="hero-subtitle">Motor Industrial LP + MILP • HiGHS • Restricciones Mixtas</p>
            <span class="v2-badge">v2.0 — Sprint 1 & 2</span>
            <p class="hero-slogan">"La exactitud es nuestra firma e innovar es nuestra naturaleza"</p>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# SIDEBAR
# ============================================================
with st.sidebar:
    st.markdown("### ⚙️ CONFIGURACIÓN DEL PROBLEMA")
    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    tipo_objetivo = st.selectbox(
        "🎯 Tipo de Objetivo",
        ["Maximizar Utilidades", "Minimizar Costos"],
        help="Selecciona si deseas maximizar ganancias o minimizar costos."
    )

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    num_productos = st.slider("📦 Número de Productos", 2, 10, 3)
    num_restricciones = st.slider("🔗 Número de Restricciones", 1, 10, 3)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # Importar desde JSON
    json_file = st.file_uploader("📂 Cargar modelo (.json)", type=["json"], key="json_upload",
                                  help="Carga un modelo guardado previamente en JSON.")
    if json_file:
        try:
            modelo_cargado = json.load(json_file)
            st.session_state['modelo_importado'] = modelo_cargado
            st.success("✅ Modelo cargado desde JSON")
        except Exception as e:
            st.error(f"Error al leer JSON: {e}")

    st.markdown("""
    <div style="background:rgba(249,168,37,0.08);border:1px solid rgba(249,168,37,0.2);
                border-radius:10px;padding:14px;margin-top:10px;">
        <p style="color:#F9A825;font-weight:700;font-size:0.8rem;margin:0;">⚡ MOTOR v2.0</p>
        <p style="color:#B0B0B0;font-size:0.75rem;margin:5px 0 0 0;">
            SciPy MILP/HiGHS<br>
            LP + MILP + Mixto<br>
            Precios Sombra (LP)
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)
    st.markdown("""
    <div style="text-align:center;padding:10px;">
        <p style="color:#707070;font-size:0.7rem;">EA Innovation & Solutions<br>Cd. Juárez, MX</p>
    </div>
    """, unsafe_allow_html=True)


# ============================================================
# TABS PRINCIPALES
# ============================================================
tab_datos, tab_importar, tab_resultados, tab_sensibilidad, tab_whatif, tab_escenarios, tab_exportar = st.tabs([
    "📋 Datos",
    "📂 Importar",
    "📊 Resultados",
    "🔬 Sensibilidad",
    "🎛️ What-If",
    "🔀 Escenarios",
    "📥 Exportar",
])


# ============================================================
# TAB: IMPORTAR DESDE EXCEL/CSV
# ============================================================
with tab_importar:
    st.markdown("""
    <div class="glass-card animate-in">
        <h3 style="color:#F9A825;margin-top:0;">📂 Importar Modelo desde Excel / CSV</h3>
        <p style="color:#B0B0B0;font-size:0.9rem;">
            Descarga la plantilla, complétala con tus datos y súbela aquí.
            Ideal para problemas con 10+ productos o restricciones.
        </p>
    </div>
    """, unsafe_allow_html=True)

    col_dl, col_up = st.columns(2)
    with col_dl:
        st.markdown("**1. Descarga la plantilla:**")
        if EXCEL_ENABLED:
            plantilla_bytes = generar_plantilla_excel()
            st.download_button(
                "📥 DESCARGAR PLANTILLA EXCEL",
                data=plantilla_bytes,
                file_name="EA_Plantilla_Modelo.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

    with col_up:
        st.markdown("**2. Sube el archivo completado:**")
        archivo_importado = st.file_uploader(
            "Archivo Excel (.xlsx)",
            type=["xlsx"],
            key="excel_upload",
            label_visibility="collapsed"
        )
        if archivo_importado and EXCEL_ENABLED:
            datos_importados = leer_excel_importado(archivo_importado)
            if datos_importados:
                st.session_state['modelo_importado'] = datos_importados
                st.success(f"✅ Importado: {len(datos_importados['nombres_productos'])} productos, "
                           f"{len(datos_importados['nombres_restricciones'])} restricciones")

    if 'modelo_importado' in st.session_state:
        m = st.session_state['modelo_importado']
        st.markdown("**Vista previa del modelo importado:**")
        df_prev = pd.DataFrame(
            m['coeficientes'],
            columns=m['nombres_productos'],
            index=m['nombres_restricciones']
        )
        df_prev.insert(0, "Tipo", m.get('tipos', ['≤'] * len(m['nombres_restricciones'])))
        df_prev["Límite"] = m['limites']
        st.dataframe(df_prev, use_container_width=True)

        st.markdown("**Productos:**")
        df_prod_prev = pd.DataFrame({
            "Producto": m['nombres_productos'],
            "Utilidad/Costo": m['margenes'],
            "Variable Entera": m.get('es_entero', [False] * len(m['nombres_productos']))
        })
        st.dataframe(df_prod_prev, use_container_width=True)


# ============================================================
# TAB: DATOS DEL PROBLEMA
# ============================================================
with tab_datos:
    # Si hay modelo importado, usar sus valores como defaults
    modelo_imp = st.session_state.get('modelo_importado', {})
    num_prod_imp = len(modelo_imp.get('nombres_productos', []))
    num_rest_imp = len(modelo_imp.get('nombres_restricciones', []))
    use_import = bool(modelo_imp)

    n_p = num_prod_imp if use_import else num_productos
    n_r = num_rest_imp if use_import else num_restricciones

    st.markdown("""
    <div class="glass-card animate-in">
        <h3 style="color:#F9A825;margin-top:0;">📦 Productos — LP Continuo o MILP Entero</h3>
        <p style="color:#B0B0B0;font-size:0.9rem;">
            Activa <b>Variable Entera</b> para productos que deben producirse en unidades completas.
            Si al menos uno es entero, se usará el solver MILP.
        </p>
    </div>
    """, unsafe_allow_html=True)

    nombres_productos, margenes, es_entero_list = [], [], []
    cols_prod = st.columns(min(n_p, 5))
    for i in range(n_p):
        col_idx = i % min(n_p, 5)
        with cols_prod[col_idx]:
            default_nombre = modelo_imp.get('nombres_productos', [f"Producto {i+1}"] * n_p)[i] if use_import else f"Producto {i+1}"
            default_margen = float(modelo_imp.get('margenes', [(i+1)*10] * n_p)[i]) if use_import else float((i+1)*10)
            default_entero = bool(modelo_imp.get('es_entero', [False] * n_p)[i]) if use_import else False

            nombre = st.text_input(f"Producto {i+1}", value=default_nombre, key=f"np_{i}")
            label_m = "💰 Utilidad/ud ($)" if tipo_objetivo == "Maximizar Utilidades" else "💲 Costo/ud ($)"
            margen = st.number_input(label_m, min_value=0.0, value=default_margen, step=0.5, key=f"mp_{i}", format="%.2f")
            entero = st.checkbox("🔢 Entero", value=default_entero, key=f"ep_{i}",
                                  help="Activar si este producto se produce en unidades enteras (sin decimales)")
            nombres_productos.append(nombre)
            margenes.append(margen)
            es_entero_list.append(entero)

    tiene_enteros = any(es_entero_list)
    modo_str = "MILP" if tiene_enteros else "LP"
    if tiene_enteros:
        productos_enteros = [nombres_productos[i] for i, e in enumerate(es_entero_list) if e]
        st.info(f"🔢 **Modo MILP activado** — Variables enteras: {', '.join(productos_enteros)}. "
                f"Los precios sombra no estarán disponibles en este modo.")

    # Función objetivo visible
    label_tipo = "MAX" if tipo_objetivo == "Maximizar Utilidades" else "MIN"
    badge_class = "badge-max" if label_tipo == "MAX" else "badge-min"
    color_obj = COLORS["gold_primary"] if label_tipo == "MAX" else COLORS["accent_blue"]
    terminos = [f"{margenes[i]:.0f}·{nombres_productos[i]}" for i in range(n_p)]
    st.markdown(f"""
    <div class="glass-card" style="border-left:3px solid {color_obj};">
        <p style="color:{COLORS['text_secondary']};font-size:0.8rem;text-transform:uppercase;letter-spacing:2px;margin-bottom:8px;">
            {'📈' if label_tipo=='MAX' else '📉'} Función Objetivo [{modo_str}]
        </p>
        <p style="color:{COLORS['white']};font-size:1.1rem;font-family:'JetBrains Mono',monospace;">
            <span class="status-badge {badge_class}">{label_tipo}</span>
            &nbsp; Z = {' + '.join(terminos)}
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

    # RESTRICCIONES
    st.markdown("""
    <div class="glass-card animate-in">
        <h3 style="color:#C62828;margin-top:0;">🔗 Restricciones — Tipos Mixtos ≤, ≥, =</h3>
        <p style="color:#B0B0B0;font-size:0.9rem;">
            Selecciona el tipo por restricción. <b>≤</b> = recurso disponible máximo.
            <b>≥</b> = demanda mínima requerida. <b>=</b> = exactamente este valor.
        </p>
    </div>
    """, unsafe_allow_html=True)

    nombres_default_rest = [
        "Horas Máquina", "Materia Prima (kg)", "Mano de Obra (hrs)",
        "Energía (kWh)", "Espacio Almacén (m²)", "Transporte (km)",
        "Presupuesto ($)", "Agua (L)", "Tiempo Horno (min)", "Empaque (unid)"
    ]

    nombres_restricciones, coeficientes_restricciones, limites, tipos_restricciones = [], [], [], []

    for j in range(n_r):
        st.markdown(f"""
        <div style="background:rgba(198,40,40,0.05);border:1px solid rgba(198,40,40,0.15);
                    border-radius:12px;padding:16px;margin-bottom:12px;">
            <p style="color:#C62828;font-weight:700;font-size:0.85rem;margin:0 0 8px 0;">
                🔗 Restricción {j+1}
            </p>
        </div>
        """, unsafe_allow_html=True)

        def_nombre_rest = modelo_imp.get('nombres_restricciones', nombres_default_rest)[j] if use_import and j < n_r else (nombres_default_rest[j] if j < len(nombres_default_rest) else f"Recurso {j+1}")
        def_tipo = modelo_imp.get('tipos', ['≤'] * n_r)[j] if use_import and j < n_r else "≤"
        def_limite = float(modelo_imp.get('limites', [100.0] * n_r)[j]) if use_import and j < n_r else 100.0

        col_name, col_tipo, col_limit = st.columns([3, 1, 1])
        with col_name:
            nombre_rest = st.text_input("Nombre del recurso", value=def_nombre_rest, key=f"nr_{j}")
        with col_tipo:
            tipo_rest = st.selectbox("Tipo", ["≤", "≥", "="], key=f"tr_{j}",
                                      index=["≤", "≥", "="].index(def_tipo) if def_tipo in ["≤", "≥", "="] else 0)
        with col_limit:
            limite = st.number_input("Límite", min_value=0.0, value=def_limite, step=1.0, key=f"lr_{j}", format="%.2f")

        nombres_restricciones.append(nombre_rest)
        tipos_restricciones.append(tipo_rest)
        limites.append(limite)

        cols_coef = st.columns(n_p)
        coefs_fila = []
        for i in range(n_p):
            def_coef = float(modelo_imp.get('coeficientes', [[1.0]*n_p]*n_r)[j][i]) if use_import and j < n_r and i < n_p else 1.0
            with cols_coef[i]:
                coef = st.number_input(
                    f"{nombres_productos[i]}",
                    min_value=0.0, value=def_coef, step=0.1,
                    key=f"c_{j}_{i}", format="%.2f",
                    help=f"Consumo de '{nombre_rest}' por unidad de '{nombres_productos[i]}'"
                )
                coefs_fila.append(coef)
        coeficientes_restricciones.append(coefs_fila)

    # Resumen del modelo
    df_modelo = pd.DataFrame(coeficientes_restricciones, columns=nombres_productos, index=nombres_restricciones)
    df_modelo.insert(0, "Tipo", tipos_restricciones)
    df_modelo["Límite"] = limites
    st.dataframe(df_modelo, use_container_width=True)

    # Validación de inputs
    warnings_val = validar_inputs(margenes, coeficientes_restricciones, nombres_productos, nombres_restricciones)
    for w in warnings_val:
        st.warning(w)

    st.markdown("<br>", unsafe_allow_html=True)
    col_btn1, col_btn2, col_btn3 = st.columns([2, 1, 2])
    with col_btn2:
        resolver = st.button("⚡ OPTIMIZAR", use_container_width=True, type="primary")


# ============================================================
# LÓGICA DE RESOLUCIÓN
# ============================================================
if resolver:
    with st.spinner(f"🔄 Ejecutando motor {modo_str} HiGHS..."):
        try:
            objetivo_str = "Maximizar" if tipo_objetivo == "Maximizar Utilidades" else "Minimizar"
            resultado = resolver_problema(
                tuple(margenes),
                tuple(tuple(row) for row in coeficientes_restricciones),
                tuple(limites),
                tuple(tipos_restricciones),
                tuple(es_entero_list),
                objetivo_str,
            )

            if resultado["success"]:
                resultado.update({
                    "nombres_productos": nombres_productos,
                    "margenes": margenes,
                    "nombres_restricciones": nombres_restricciones,
                    "limites": limites,
                    "tipos_restricciones": tipos_restricciones,
                    "coeficientes": coeficientes_restricciones,
                    "tipo_objetivo": tipo_objetivo,
                    "es_entero_list": es_entero_list,
                    "timestamp": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                })
                st.session_state["resultado_simplex"] = resultado
                st.success(f"✅ Solución {modo_str} óptima encontrada — ${resultado['valor_objetivo']:,.2f}. Navega a los resultados.")
                st.balloons()
            else:
                st.session_state["resultado_simplex"] = resultado
                status = resultado.get("status", "?")
                msg = resultado.get("mensaje", "Error desconocido")
                st.error(f"❌ Estado {status}: {msg}")

        except Exception as e:
            st.error(f"⚠️ Error inesperado: {str(e)}")
            st.session_state["resultado_simplex"] = {"success": False, "mensaje": str(e)}


# ============================================================
# TAB: RESULTADOS ÓPTIMOS
# ============================================================
with tab_resultados:
    if "resultado_simplex" not in st.session_state or not st.session_state["resultado_simplex"].get("success"):
        st.markdown("""
        <div class="glass-card" style="text-align:center;padding:60px;">
            <p class="pulse-icon" style="font-size:60px;">⏳</p>
            <h2 style="color:#F5F5F5;">Esperando Optimización</h2>
            <p style="color:#B0B0B0;">Configura los datos y presiona <strong style="color:#C62828;">⚡ OPTIMIZAR</strong>.</p>
        </div>
        """, unsafe_allow_html=True)
    else:
        res = st.session_state["resultado_simplex"]
        label_valor = "Utilidad Máxima" if res["tipo_objetivo"] == "Maximizar Utilidades" else "Costo Mínimo"
        color_kpi = COLORS["gold_primary"] if "Maximizar" in res["tipo_objetivo"] else COLORS["accent_blue"]
        modo_label = "MILP" if res.get("tiene_enteros") else "LP"

        st.markdown(f"""
        <div class="glass-card animate-in" style="text-align:center;border-top:3px solid {color_kpi};">
            <p style="color:{COLORS['text_secondary']};font-size:0.85rem;text-transform:uppercase;letter-spacing:3px;margin-bottom:8px;">
                {'📈' if 'Maximizar' in res['tipo_objetivo'] else '📉'} {label_valor}
                &nbsp;<span class="v2-badge">{modo_label}</span>
            </p>
            <p style="font-size:3.2rem;font-weight:900;color:{color_kpi};font-family:'JetBrains Mono',monospace;margin:0;line-height:1;">
                ${res['valor_objetivo']:,.2f}
            </p>
            <p style="color:{COLORS['text_muted']};font-size:0.8rem;margin-top:10px;">
                Motor: SciPy {modo_label}/HiGHS • Ejecutado: {res['timestamp']}
            </p>
        </div>
        """, unsafe_allow_html=True)

        contribuciones = [res["variables"][i] * res["margenes"][i] for i in range(len(res["variables"]))]
        cols_kpi = st.columns(len(res["nombres_productos"]))
        for i, (nombre, cantidad) in enumerate(zip(res["nombres_productos"], res["variables"])):
            with cols_kpi[i]:
                tipo_var = "🔢 Entero" if res["es_entero_list"][i] else "📊 Continuo"
                st.metric(label=f"📦 {nombre} ({tipo_var})",
                          value=f"{cantidad:,.2f} uds",
                          delta=f"${contribuciones[i]:,.2f}")

        st.markdown('<div class="section-divider"></div>', unsafe_allow_html=True)

        col_c1, col_c2 = st.columns(2)
        with col_c1:
            df_prod = pd.DataFrame({
                "Producto": res["nombres_productos"],
                "Cantidad": res["variables"],
                "Contribución": contribuciones
            })
            colors_bar = ["#CE93D8" if res["es_entero_list"][i] else "#C62828" for i in range(len(res["variables"]))]
            fig_b = go.Figure()
            fig_b.add_trace(go.Bar(
                x=df_prod["Producto"], y=df_prod["Cantidad"],
                marker_color=colors_bar,
                text=[f"{v:,.2f}" for v in df_prod["Cantidad"]],
                textposition="outside",
                textfont=dict(color="#F5F5F5", size=13)
            ))
            fig_b.update_layout(
                title=dict(text="Unidades Óptimas por Producto<br><sup>Morado=Entero | Rojo=Continuo</sup>", font=dict(color="#FFFFFF", size=16)),
                template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(26,26,26,0.8)",
                height=400, showlegend=False
            )
            st.plotly_chart(fig_b, use_container_width=True)

        with col_c2:
            colors_pie = ["#C62828","#F9A825","#E53935","#FDD835","#42A5F5","#EF5350","#FFB74D","#66BB6A","#AB47BC","#26A69A"]
            fig_dona = go.Figure(data=[go.Pie(
                labels=df_prod["Producto"], values=df_prod["Contribución"], hole=0.55,
                marker=dict(colors=colors_pie[:len(df_prod)], line=dict(color="#0D0D0D", width=2)),
                textinfo="label+percent",
                hovertemplate="<b>%{label}</b><br>Contribución: $%{value:,.2f}<br>%{percent}<extra></extra>"
            )])
            fig_dona.update_layout(
                title=dict(text="Distribución de Contribución", font=dict(color="#FFFFFF", size=16)),
                template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)",
                height=400, margin=dict(t=60, b=60),
                annotations=[dict(text=f"<b>${res['valor_objetivo']:,.0f}</b>", x=0.5, y=0.5,
                                  font_size=20, font_color=COLORS["gold_primary"], showarrow=False)]
            )
            st.plotly_chart(fig_dona, use_container_width=True)

        # Altair comparativo
        df_alt = pd.DataFrame({
            "Producto": res["nombres_productos"],
            "Cantidad Producida": res["variables"],
            "Margen Unitario ($)": res["margenes"],
            "Contribución Total ($)": contribuciones,
        })
        chart_alt = alt.Chart(df_alt).mark_bar(cornerRadiusTopRight=8, cornerRadiusBottomRight=8).encode(
            x=alt.X("Contribución Total ($):Q", title="Contribución Total ($)"),
            y=alt.Y("Producto:N", sort="-x", title=""),
            color=alt.Color("Contribución Total ($):Q", scale=alt.Scale(scheme="redyellowgreen"), legend=None),
            tooltip=[alt.Tooltip("Producto:N"), alt.Tooltip("Cantidad Producida:Q", format=",.2f"),
                     alt.Tooltip("Margen Unitario ($):Q", format="$,.2f"),
                     alt.Tooltip("Contribución Total ($):Q", format="$,.2f")]
        ).properties(
            height=max(200, len(df_alt) * 60),
            title=alt.Title(text="Contribución por Producto al Objetivo", color="#FFFFFF", fontSize=15)
        ).configure(background="rgba(26,26,26,0.8)").configure_axis(
            labelColor="#E0E0E0", titleColor="#F9A825", gridColor="rgba(255,255,255,0.06)"
        )
        st.altair_chart(chart_alt, use_container_width=True)


# ============================================================
# TAB: ANÁLISIS DE SENSIBILIDAD (con precios sombra)
# ============================================================
with tab_sensibilidad:
    if "resultado_simplex" not in st.session_state or not st.session_state["resultado_simplex"].get("success"):
        st.markdown("""
        <div class="glass-card" style="text-align:center;padding:60px;">
            <p class="pulse-icon" style="font-size:60px;">🔬</p>
            <h2 style="color:#F5F5F5;">Análisis Pendiente</h2>
        </div>
        """, unsafe_allow_html=True)
    else:
        res = st.session_state["resultado_simplex"]
        tiene_enteros_res = res.get("tiene_enteros", False)

        st.markdown(f"""
        <div class="glass-card animate-in">
            <h3 style="color:#C62828;margin-top:0;">🔬 Análisis de Sensibilidad — Precios Sombra</h3>
            <p style="color:#B0B0B0;font-size:0.9rem;">
                {'⚠️ Modo MILP: los precios sombra no aplican. Desactiva variables enteras para obtener análisis dual completo.' if tiene_enteros_res else
                '✅ Modo LP: precios sombra disponibles. Indica cuánto mejora el objetivo por cada unidad adicional del recurso.'}
            </p>
        </div>
        """, unsafe_allow_html=True)

        restricciones_activas = 0
        col_h = st.columns([2, 1, 1, 1, 1, 1.5, 1.5])
        headers_s = ["🔗 Recurso", "Tipo", "📦 Disponible", "⚙️ Consumido", "📐 Holgura", "📊 % Util.", "💰 P.Sombra"]
        for c, h in zip(col_h, headers_s):
            c.markdown(f"**{h}**")
        st.markdown("---")

        for j in range(len(res["nombres_restricciones"])):
            disp = res["limites"][j]
            cons = res["consumo"][j]
            hol = res["holguras"][j]
            tipo_r = res.get("tipos_restricciones", ["≤"] * len(res["limites"]))[j]
            dual = res["duals"][j] if not tiene_enteros_res else None
            pct = (cons / disp * 100) if disp > 0 else 0
            activa = hol < 0.01

            if activa:
                restricciones_activas += 1

            c1, c2, c3, c4, c5, c6, c7 = st.columns([2, 1, 1, 1, 1, 1.5, 1.5])
            c1.markdown(f"**{res['nombres_restricciones'][j]}**")
            c2.markdown(f"`{tipo_r}`")
            c3.markdown(f"`{disp:,.2f}`")
            c4.markdown(f"`{cons:,.2f}`")
            color_hol = "🔴" if activa else "🟢"
            c5.markdown(f"{color_hol} `{hol:,.2f}`")
            c6.progress(min(pct / 100.0, 1.0), text=f"{pct:.1f}%")
            if dual is not None:
                shadow_color = "🟡" if abs(dual) > 0.001 else "⚪"
                c7.markdown(f"{shadow_color} `{dual:,.4f}`")
            else:
                c7.markdown("*N/A (MILP)*")

        if not tiene_enteros_res:
            st.info("💰 **Precio Sombra:** valor del objetivo que se gana/pierde por cada unidad adicional del recurso. "
                    "Un precio sombra de 5.0 significa que conseguir 1 unidad más de ese recurso aumentaría la utilidad en $5.00.")

        total_rest = len(res["nombres_restricciones"])
        pct_cuellos = (restricciones_activas / total_rest * 100) if total_rest > 0 else 0

        st.markdown(f"""
        <div class="kpi-container" style="margin-top:24px;">
            <div class="kpi-card red"><p class="kpi-label">⛔ Cuellos de Botella</p>
                <p class="kpi-value">{restricciones_activas}</p><p class="kpi-unit">activas</p></div>
            <div class="kpi-card gold"><p class="kpi-label">✅ Con Holgura</p>
                <p class="kpi-value">{total_rest - restricciones_activas}</p><p class="kpi-unit">disponibles</p></div>
            <div class="kpi-card blue"><p class="kpi-label">📊 Saturación Media</p>
                <p class="kpi-value">{np.mean([(res['consumo'][j]/res['limites'][j]*100) if res['limites'][j]>0 else 0 for j in range(len(res['limites']))]):.1f}%</p>
                <p class="kpi-unit">utilización</p></div>
            <div class="kpi-card success"><p class="kpi-label">🏭 Presión de Planta</p>
                <p class="kpi-value">{pct_cuellos:.0f}%</p><p class="kpi-unit">saturadas</p></div>
        </div>
        """, unsafe_allow_html=True)

        # Gráfico Consumo vs Límite
        df_rec = pd.DataFrame({
            "Recurso": res["nombres_restricciones"],
            "Disponible": res["limites"],
            "Consumido": res["consumo"],
            "Holgura": res["holguras"],
            "P. Sombra": res["duals"] if not tiene_enteros_res else [0] * len(res["limites"])
        })
        fig_r = go.Figure()
        fig_r.add_trace(go.Bar(name="Consumido", x=df_rec["Recurso"], y=df_rec["Consumido"],
                               marker_color="#C62828", text=[f"{v:,.1f}" for v in df_rec["Consumido"]], textposition="inside", textfont=dict(color="#FFF", size=13)))
        fig_r.add_trace(go.Bar(name="Holgura", x=df_rec["Recurso"], y=df_rec["Holgura"],
                               marker_color="#F9A825", text=[f"{v:,.1f}" for v in df_rec["Holgura"]], textposition="inside", textfont=dict(color="#000", size=13)))
        if not tiene_enteros_res:
            fig_r.add_trace(go.Scatter(name="Precio Sombra", x=df_rec["Recurso"], y=df_rec["P. Sombra"],
                                       mode="markers+lines", yaxis="y2",
                                       marker=dict(color="#42A5F5", size=10, symbol="diamond"),
                                       line=dict(color="#42A5F5", width=2, dash="dot")))
        fig_r.update_layout(
            barmode="stack", template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(26,26,26,0.8)",
            yaxis=dict(title="Cantidad", title_font=dict(color="#F9A825")),
            yaxis2=dict(title="Precio Sombra", overlaying="y", side="right", title_font=dict(color="#42A5F5")) if not tiene_enteros_res else {},
            legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
            height=450, margin=dict(t=60, b=50)
        )
        st.plotly_chart(fig_r, use_container_width=True)

        # Interpretación
        activas_names = [res["nombres_restricciones"][j] for j in range(len(res["holguras"])) if res["holguras"][j] < 0.01]
        holgadas_names = [res["nombres_restricciones"][j] for j in range(len(res["holguras"])) if res["holguras"][j] >= 0.01]
        if activas_names:
            st.warning(f"⚠️ **Cuellos de Botella:** {', '.join(activas_names)} — utilizados al 100%. "
                       "Incrementar su capacidad mejoraría el objetivo directamente.")
        if holgadas_names:
            max_h = int(np.argmax(res["holguras"]))
            st.success(f"✅ **Recursos con Excedente:** {', '.join(holgadas_names)}. "
                       f"Mayor holgura: **{res['nombres_restricciones'][max_h]}** ({res['holguras'][max_h]:,.2f} unidades sobrantes).")


# ============================================================
# TAB: WHAT-IF INTERACTIVO
# ============================================================
with tab_whatif:
    if "resultado_simplex" not in st.session_state or not st.session_state["resultado_simplex"].get("success"):
        st.markdown("""
        <div class="glass-card" style="text-align:center;padding:60px;">
            <p class="pulse-icon" style="font-size:60px;">🎛️</p>
            <h2 style="color:#F5F5F5;">Optimiza primero</h2>
        </div>
        """, unsafe_allow_html=True)
    else:
        res = st.session_state["resultado_simplex"]
        st.markdown("""
        <div class="glass-card animate-in">
            <h3 style="color:#F9A825;margin-top:0;">🎛️ Análisis What-If — ¿Qué pasa si...?</h3>
            <p style="color:#B0B0B0;font-size:0.9rem;">
                Ajusta los límites de recursos en tiempo real para ver cómo impactan el valor objetivo.
                El solver recalcula automáticamente al mover los sliders.
            </p>
        </div>
        """, unsafe_allow_html=True)

        limites_whatif = []
        col_sliders = st.columns(min(len(res["nombres_restricciones"]), 3))
        for j, (nombre, limite_orig) in enumerate(zip(res["nombres_restricciones"], res["limites"])):
            with col_sliders[j % 3]:
                tipo_r = res.get("tipos_restricciones", ["≤"] * len(res["limites"]))[j]
                new_lim = st.slider(
                    f"{nombre} [{tipo_r}]",
                    min_value=max(0.0, limite_orig * 0.2),
                    max_value=limite_orig * 3.0,
                    value=float(limite_orig),
                    step=max(1.0, limite_orig * 0.05),
                    key=f"wi_{j}",
                    format="%.1f"
                )
                limites_whatif.append(new_lim)

        # Recalcular con nuevos límites
        try:
            obj_str = "Maximizar" if res["tipo_objetivo"] == "Maximizar Utilidades" else "Minimizar"
            res_wi = resolver_problema(
                tuple(res["margenes"]),
                tuple(tuple(row) for row in res["coeficientes"]),
                tuple(limites_whatif),
                tuple(res.get("tipos_restricciones", ["≤"] * len(res["limites"]))),
                tuple(res.get("es_entero_list", [False] * len(res["margenes"]))),
                obj_str,
            )
            if res_wi["success"]:
                delta_val = res_wi["valor_objetivo"] - res["valor_objetivo"]
                color_delta = COLORS["success"] if delta_val >= 0 else COLORS["error"]
                arrow = "▲" if delta_val >= 0 else "▼"
                st.markdown(f"""
                <div class="glass-card" style="text-align:center;border-top:3px solid {color_delta};">
                    <p style="color:{COLORS['text_secondary']};font-size:0.85rem;text-transform:uppercase;letter-spacing:2px;">
                        Valor Objetivo con nuevos límites
                    </p>
                    <p style="font-size:2.8rem;font-weight:900;color:{COLORS['gold_primary']};font-family:'JetBrains Mono',monospace;margin:0;">
                        ${res_wi['valor_objetivo']:,.2f}
                    </p>
                    <p style="font-size:1.2rem;color:{color_delta};font-weight:700;margin-top:8px;">
                        {arrow} ${abs(delta_val):,.2f} vs. solución base
                    </p>
                </div>
                """, unsafe_allow_html=True)

                # Comparación de producción
                df_wi = pd.DataFrame({
                    "Producto": res["nombres_productos"],
                    "Base": res["variables"],
                    "What-If": res_wi["variables"],
                })
                fig_wi = go.Figure()
                fig_wi.add_trace(go.Bar(name="Solución Base", x=df_wi["Producto"], y=df_wi["Base"], marker_color="#C62828"))
                fig_wi.add_trace(go.Bar(name="What-If", x=df_wi["Producto"], y=df_wi["What-If"], marker_color="#F9A825"))
                fig_wi.update_layout(
                    barmode="group", template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(26,26,26,0.8)",
                    title="Comparación: Base vs. What-If",
                    legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
                    height=380
                )
                st.plotly_chart(fig_wi, use_container_width=True)
            else:
                st.error(f"❌ {res_wi['mensaje']}")
        except Exception as e:
            st.error(f"Error en What-If: {e}")


# ============================================================
# TAB: COMPARADOR DE ESCENARIOS
# ============================================================
with tab_escenarios:
    if "escenarios" not in st.session_state:
        st.session_state["escenarios"] = []

    st.markdown("""
    <div class="glass-card animate-in">
        <h3 style="color:#F9A825;margin-top:0;">🔀 Comparador de Escenarios</h3>
        <p style="color:#B0B0B0;font-size:0.9rem;">
            Guarda hasta 5 corridas y compáralas lado a lado.
            Útil para analizar cambios en precios, demandas o capacidades.
        </p>
    </div>
    """, unsafe_allow_html=True)

    col_save, col_clear = st.columns([3, 1])
    with col_save:
        nombre_escenario = st.text_input("Nombre del escenario", value=f"Escenario {len(st.session_state['escenarios']) + 1}", key="esc_nombre")
    with col_clear:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("🗑️ Limpiar Todo", use_container_width=True):
            st.session_state["escenarios"] = []
            st.rerun()

    col_b1, col_b2 = st.columns(2)
    with col_b1:
        if st.button("💾 GUARDAR ESCENARIO ACTUAL", use_container_width=True):
            if "resultado_simplex" in st.session_state and st.session_state["resultado_simplex"].get("success"):
                res_now = st.session_state["resultado_simplex"]
                esc = {
                    "nombre": nombre_escenario,
                    "tipo_objetivo": res_now["tipo_objetivo"],
                    "valor_objetivo": res_now["valor_objetivo"],
                    "variables": res_now["variables"],
                    "nombres_productos": res_now["nombres_productos"],
                    "nombres_restricciones": res_now["nombres_restricciones"],
                    "consumo": res_now["consumo"],
                    "limites": res_now["limites"],
                    "tiene_enteros": res_now.get("tiene_enteros", False),
                    "timestamp": res_now["timestamp"],
                }
                if len(st.session_state["escenarios"]) >= 5:
                    st.warning("Máximo 5 escenarios. Elimina uno para agregar más.")
                else:
                    st.session_state["escenarios"].append(esc)
                    st.success(f"✅ Escenario '{nombre_escenario}' guardado.")
                    st.rerun()
            else:
                st.warning("No hay resultado activo. Ejecuta la optimización primero.")

    if st.session_state["escenarios"]:
        escenarios = st.session_state["escenarios"]
        # Tabla comparativa de valores objetivo
        df_esc = pd.DataFrame([{
            "Escenario": e["nombre"],
            "Objetivo": e["tipo_objetivo"],
            "Valor Óptimo": f"${e['valor_objetivo']:,.2f}",
            "Modo": "MILP" if e.get("tiene_enteros") else "LP",
            "Timestamp": e["timestamp"],
        } for e in escenarios])
        st.dataframe(df_esc, use_container_width=True, hide_index=True)

        # Gráfico comparativo de valores objetivo
        fig_esc = go.Figure(go.Bar(
            x=[e["nombre"] for e in escenarios],
            y=[e["valor_objetivo"] for e in escenarios],
            marker_color=["#C62828", "#F9A825", "#42A5F5", "#00E676", "#AB47BC"][:len(escenarios)],
            text=[f"${e['valor_objetivo']:,.2f}" for e in escenarios],
            textposition="outside"
        ))
        fig_esc.update_layout(
            title="Comparación de Valores Objetivo por Escenario",
            template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(26,26,26,0.8)",
            height=380, showlegend=False
        )
        st.plotly_chart(fig_esc, use_container_width=True)

        # Comparación de producción por escenario
        if len(escenarios) >= 2:
            prod_names = escenarios[0]["nombres_productos"]
            fig_prod = go.Figure()
            colores_esc = ["#C62828", "#F9A825", "#42A5F5", "#00E676", "#AB47BC"]
            for k, esc in enumerate(escenarios):
                fig_prod.add_trace(go.Bar(
                    name=esc["nombre"],
                    x=esc["nombres_productos"],
                    y=esc["variables"],
                    marker_color=colores_esc[k % len(colores_esc)]
                ))
            fig_prod.update_layout(
                barmode="group", template="plotly_dark", paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(26,26,26,0.8)",
                title="Producción Óptima por Escenario",
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="center", x=0.5),
                height=420
            )
            st.plotly_chart(fig_prod, use_container_width=True)
    else:
        st.markdown("""
        <div class="glass-card" style="text-align:center;padding:40px;">
            <p style="color:#707070;font-size:1rem;">No hay escenarios guardados. Ejecuta la optimización y guarda escenarios para compararlos.</p>
        </div>
        """, unsafe_allow_html=True)


# ============================================================
# TAB: EXPORTAR
# ============================================================
with tab_exportar:
    if "resultado_simplex" not in st.session_state or not st.session_state["resultado_simplex"].get("success"):
        st.markdown("""
        <div class="glass-card" style="text-align:center;padding:60px;">
            <p class="pulse-icon" style="font-size:60px;">📥</p>
            <h2 style="color:#F5F5F5;">Sin Datos para Exportar</h2>
        </div>
        """, unsafe_allow_html=True)
    else:
        res = st.session_state["resultado_simplex"]
        contribuciones_exp = [res["variables"][i] * res["margenes"][i] for i in range(len(res["variables"]))]
        datos_reporte = {
            "objetivo": res["tipo_objetivo"],
            "num_productos": len(res["nombres_productos"]),
            "num_restricciones": len(res["nombres_restricciones"]),
            "valor_objetivo": res["valor_objetivo"],
            "tiene_enteros": res.get("tiene_enteros", False),
            "productos": [{
                "nombre": res["nombres_productos"][i],
                "cantidad": res["variables"][i],
                "contribucion": contribuciones_exp[i],
                "es_entero": res.get("es_entero_list", [False]*len(res["variables"]))[i],
            } for i in range(len(res["variables"]))],
            "recursos": [{
                "nombre": res["nombres_restricciones"][j],
                "tipo": res.get("tipos_restricciones", ["≤"]*len(res["limites"]))[j],
                "disponible": res["limites"][j],
                "consumido": res["consumo"][j],
                "holgura": res["holguras"][j],
                "dual": res["duals"][j],
            } for j in range(len(res["limites"]))],
            "escenarios": st.session_state.get("escenarios", []),
        }

        st.markdown("""
        <div class="glass-card animate-in">
            <h3 style="color:#F9A825;margin-top:0;">📥 Centro de Exportación v2.0</h3>
        </div>
        """, unsafe_allow_html=True)

        col_pdf, col_excel, col_json = st.columns(3)

        with col_pdf:
            st.markdown("""
            <div class="glass-card" style="text-align:center;border-top:3px solid #C62828;">
                <p style="font-size:40px;margin-bottom:8px;">📄</p>
                <h3 style="color:#C62828;margin:0;">Reporte PDF</h3>
                <p style="color:#B0B0B0;font-size:0.8rem;margin-top:8px;">
                    Con precios sombra y tipo MILP/LP
                </p>
            </div>
            """, unsafe_allow_html=True)
            if PDF_ENABLED:
                pdf_b = generar_pdf_reporte(datos_reporte)
                if pdf_b:
                    st.download_button(
                        "📄 DESCARGAR PDF", data=pdf_b,
                        file_name=f"EA_Simplex_v2_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                        mime="application/pdf", use_container_width=True
                    )
            else:
                st.warning("Instala `fpdf2` para PDF.")

        with col_excel:
            st.markdown("""
            <div class="glass-card" style="text-align:center;border-top:3px solid #F9A825;">
                <p style="font-size:40px;margin-bottom:8px;">📊</p>
                <h3 style="color:#F9A825;margin:0;">Reporte Excel</h3>
                <p style="color:#B0B0B0;font-size:0.8rem;margin-top:8px;">
                    Con escenarios y precios sombra
                </p>
            </div>
            """, unsafe_allow_html=True)
            if EXCEL_ENABLED:
                excel_b = generar_excel_reporte(datos_reporte)
                if excel_b:
                    st.download_button(
                        "📊 DESCARGAR EXCEL", data=excel_b,
                        file_name=f"EA_Simplex_v2_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )

        with col_json:
            st.markdown("""
            <div class="glass-card" style="text-align:center;border-top:3px solid #42A5F5;">
                <p style="font-size:40px;margin-bottom:8px;">🗄️</p>
                <h3 style="color:#42A5F5;margin:0;">Modelo JSON</h3>
                <p style="color:#B0B0B0;font-size:0.8rem;margin-top:8px;">
                    Guarda y reutiliza el modelo completo
                </p>
            </div>
            """, unsafe_allow_html=True)
            modelo_json = {
                "version": "2.0",
                "tipo_objetivo": res["tipo_objetivo"],
                "nombres_productos": res["nombres_productos"],
                "margenes": res["margenes"],
                "es_entero": res.get("es_entero_list", [False]*len(res["margenes"])),
                "nombres_restricciones": res["nombres_restricciones"],
                "tipos": res.get("tipos_restricciones", ["≤"]*len(res["limites"])),
                "limites": res["limites"],
                "coeficientes": res["coeficientes"],
                "resultado": {
                    "variables": res["variables"],
                    "valor_objetivo": res["valor_objetivo"],
                    "timestamp": res["timestamp"],
                }
            }
            st.download_button(
                "🗄️ GUARDAR MODELO JSON",
                data=json.dumps(modelo_json, indent=2, ensure_ascii=False),
                file_name=f"EA_Modelo_{datetime.now().strftime('%Y%m%d_%H%M')}.json",
                mime="application/json",
                use_container_width=True
            )


# ============================================================
# FOOTER
# ============================================================
st.markdown("""
<div class="footer-premium">
    <p class="footer-slogan">"La exactitud es nuestra firma e innovar es nuestra naturaleza"</p>
    <p class="footer-credits">
        🏭 EA Simplex Production Optimizer v2.0 &nbsp;|&nbsp;
        🐍 Ing. Maestro Erik Armenta &nbsp;|&nbsp;
        ⚡ SciPy MILP/HiGHS &nbsp;|&nbsp;
        📍 EA Innovation & Solutions — Cd. Juárez, MX<br>
        <span style="color:#C62828;">●</span> MILP
        <span style="color:#F9A825;">●</span> Restricciones Mixtas
        <span style="color:#42A5F5;">●</span> Precios Sombra
        <span style="color:#00E676;">●</span> Comparador Escenarios
        <span style="color:#AB47BC;">●</span> What-If
        <span style="color:#FF7043;">●</span> JSON Persistence
    </p>
</div>
""", unsafe_allow_html=True)
